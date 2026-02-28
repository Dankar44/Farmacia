"""
Scraper de FarmaciasVazquez - Usa la API de Doofinder.

FarmaciasVazquez usa Doofinder como motor de búsqueda.
Las peticiones van directamente a los servidores de Doofinder,
por lo que es rapidísimo y no hay bloqueos anti-bot.

Credenciales Doofinder (descubiertas por ingeniería inversa):
  - HashID: b8385fd3e2f32aadf43c359fb6791646
  - Endpoint: https://eu1-search.doofinder.com/5/search
  - Límite por petición: 100 resultados
  - Límite por búsqueda general: 1000 resultados

Para extraer el catálogo completo (~25k productos), el scraper itera 
por letras del abecedario y números para evitar el límite de 1000 resultados 
por consulta.

Uso:
    python main.py vazquez                     # Todos los productos (~25k, ~2 min)
    python main.py vazquez --limit 100         # Solo 100 productos (test)
    python main.py vazquez --export            # Solo exportar datos a Excel
"""

import os
import sys
import time
import re
import xml.etree.ElementTree as ET
import logging
import argparse
import requests
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from sqlalchemy.orm import sessionmaker

# Directorio raíz del proyecto
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
from db_models import get_engine, Producto, Precio, Base

# ============================================================
# LOGGING
# ============================================================
LOG_DIR = os.path.join(PROJECT_ROOT, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, 'scraper_vazquez.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ============================================================
# CONFIGURACIÓN DE DOOFINDER PARA VAZQUEZ
# ============================================================
DOOFINDER_HASHID = "b8385fd3e2f32aadf43c359fb6791646"
DOOFINDER_URL = "https://eu1-search.doofinder.com/5/search"

FARMACIA_NOMBRE = "FarmaciasVazquez"
BASE_URL = "https://www.farmavazquez.com"

# Delay entre peticiones a la API (segundos)
DELAY = 0.2

# Límite técnico de Doofinder
HITS_PER_PAGE = 100
MAX_RESULTS_PER_QUERY = 1000

# Cuántos productos guardar por lote antes de commit
BATCH_SIZE = 100


# ============================================================
# CONSULTA A DOOFINDER
# ============================================================
def consultar_doofinder(query="", page=1):
    """Hace una consulta a la API de Doofinder de FarmaciasVazquez."""
    params = {
        "hashid": DOOFINDER_HASHID,
        "query": query,
        "page": page,
        "rpp": HITS_PER_PAGE
    }

    headers = {
        "Origin": "https://www.farmavazquez.com",
        "Referer": "https://www.farmavazquez.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }

    try:
        response = requests.get(DOOFINDER_URL, params=params, headers=headers, timeout=15)
    except requests.exceptions.RequestException as e:
        logger.error(f"Error de conexión con Doofinder: {e}")
        return None

    if response.status_code != 200:
        logger.error(f"Error API Doofinder: HTTP {response.status_code}")
        return None

    return response.json()


def consultar_doofinder_batch(ids_list):
    """Consulta un lote exacto de IDs en Doofinder."""
    params = {
        "hashid": DOOFINDER_HASHID,
        "query": "",
        "filter[id]": ids_list,
        "page": 1,
        "rpp": len(ids_list)
    }

    headers = {
        "Origin": "https://www.farmavazquez.com",
        "Referer": "https://www.farmavazquez.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }

    try:
        response = requests.get(DOOFINDER_URL, params=params, headers=headers, timeout=15)
        if response.status_code == 200:
            return response.json()
    except Exception as e:
        logger.error(f"Error de conexión con Doofinder Batch: {e}")
    return None


# ============================================================
# EXTRACCIÓN DE DATOS
# ============================================================
def extraer_datos_producto(hit):
    """Extrae datos relevantes de un producto de Doofinder."""
    nombre = hit.get("title", "Sin nombre")
    url = hit.get("link", "")
    sku = hit.get("dfid", "")
    
    # Doofinder de Vázquez tiene stock como int (cantidad)
    stock_q = hit.get("stock_quantity", 0)
    en_stock = stock_q > 0

    # Marca
    brand = hit.get("brand", "")
    marca = brand if isinstance(brand, str) else ""

    # Categorías
    categorias = hit.get("categories", [])
    categoria = ""
    if categorias and isinstance(categorias, list):
        # Usar la última categoría que suele ser la más específica
        categoria = categorias[-1] if isinstance(categorias[-1], str) else str(categorias[-1])
        
    if not categoria:
        categoria = hit.get("main_category", "")

    # Precio
    precio = None
    precio_original = None

    try:
        p_val = hit.get("price")
        p_sale = hit.get("best_price") or hit.get("sale_price")
        
        precio_api = Decimal(str(p_val)) if p_val is not None else None
        precio_sale_api = Decimal(str(p_sale)) if p_sale is not None else None

        if precio_api and precio_sale_api and precio_sale_api < precio_api:
            # Hay descuento
            precio = precio_sale_api
            precio_original = round(precio_api, 2)
        elif precio_sale_api:
            precio = precio_sale_api
        elif precio_api:
            precio = precio_api
            
        if precio:
            precio = round(precio, 2)
                
    except (InvalidOperation, ValueError, TypeError):
        pass

    # URL completa
    if url and not url.startswith("http"):
        url = f"{BASE_URL}{url}"

    return {
        "nombre": nombre,
        "url": url,
        "sku": sku,
        "precio": precio,
        "precio_original": precio_original,
        "en_stock": en_stock,
        "categoria": categoria,
        "marca": marca,
    }


# ============================================================
# EXTRACCIÓN SITEMAP Y SCRAPING
# ============================================================
def obtener_ids_sitemap():
    """Descarga el sitemap de Vazquez y extrae todos los IDs de productos."""
    logger.info("Descargando índice de sitemaps...")
    headers = {'User-Agent': 'Mozilla/5.0'}
    index_url = "https://www.farmavazquez.com/gsitemap/2_index_sitemap.xml"
    
    ids_productos = set()
    
    try:
        resp = requests.get(index_url, headers=headers, timeout=15)
        if resp.status_code != 200:
            logger.error(f"Error al descargar sitemap index: HTTP {resp.status_code}")
            return list(ids_productos)
            
        root = ET.fromstring(resp.content)
        sitemap_urls = []
        for sitemap in root.findall('{http://www.sitemaps.org/schemas/sitemap/0.9}sitemap'):
            loc = sitemap.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc')
            if loc is not None:
                sitemap_urls.append(loc.text)
                
        logger.info(f"Encontrados {len(sitemap_urls)} sitemaps secundarios.")
        
        id_pattern = re.compile(r'-(\d+)\.html$')
        
        for s_url in sitemap_urls:
            logger.info(f"Procesando sitemap: {s_url}")
            try:
                s_resp = requests.get(s_url, headers=headers, timeout=15)
                s_root = ET.fromstring(s_resp.content)
                for url_node in s_root.findall('{http://www.sitemaps.org/schemas/sitemap/0.9}url'):
                    loc = url_node.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc')
                    if loc is not None and loc.text:
                        match = id_pattern.search(loc.text)
                        if match:
                            ids_productos.add(match.group(1))
            except Exception as e:
                logger.error(f"Error parseando sitemap {s_url}: {e}")
                
    except Exception as e:
        logger.error(f"Error descargando el sitemap principal: {e}")
        
    logger.info(f"Se han extraído {len(ids_productos)} IDs únicos del catálogo de sitemaps.")
    return list(ids_productos)

def ejecutar_scraping(db, limit=0):
    """Extrae los productos consultando Doofinder por lotes de IDs."""
    
    hoy = datetime.now(timezone.utc)
    exitos = 0
    errores = 0

    todas_ids = obtener_ids_sitemap()
    if not todas_ids:
        logger.error("No se encontraron IDs en los sitemaps. Abortando.")
        return

    if limit > 0:
        todas_ids = todas_ids[:limit]
        
    # Agrupar IDs en lotes de 100
    chunks = [todas_ids[i:i + 100] for i in range(0, len(todas_ids), 100)]
    logger.info(f"Iniciando extracción por lotes: {len(chunks)} peticiones a Doofinder.")

    for i, chunk in enumerate(chunks, 1):
        if i > 1:
            time.sleep(DELAY)
            
        logger.info(f"Procesando lote {i}/{len(chunks)} ({len(chunk)} productos)...")
        data = consultar_doofinder_batch(chunk)
        if not data or 'results' not in data:
            errores += len(chunk)
            continue
            
        resultados = data.get("results", [])
        
        for hit in resultados:
            datos = extraer_datos_producto(hit)
            
            if not datos["url"] or datos["precio"] is None:
                errores += 1
                continue
                
            # Guardar en DB
            producto = db.query(Producto).filter_by(url=datos["url"]).first()

            if not producto:
                producto = Producto(
                    nombre=datos["nombre"],
                    url=datos["url"],
                    farmacia=FARMACIA_NOMBRE,
                    categoria=datos["categoria"],
                    ean=datos["sku"],
                )
                db.add(producto)
                db.flush()
            else:
                producto.nombre = datos["nombre"]
                producto.categoria = datos["categoria"]
                if datos["sku"]:
                    producto.ean = datos["sku"]

            precio_record = Precio(
                producto_id=producto.id,
                precio=datos["precio"],
                precio_original=datos["precio_original"],
                en_stock=datos["en_stock"],
                fecha_captura=hoy,
            )
            db.add(precio_record)
            exitos += 1

        db.commit()

    logger.info("=" * 60)
    logger.info("RESUMEN FINAL")
    logger.info("=" * 60)
    logger.info(f"Productos extraídos hoy: {exitos}")
    logger.info(f"Productos sin precio/URL (o erróneos): {errores}")
    logger.info(f"Total histórico en DB: {db.query(Producto).filter_by(farmacia=FARMACIA_NOMBRE).count()}")


# ============================================================
# EXPORTAR A EXCEL
# ============================================================
def exportar_a_excel(db, filename=None):
    if filename is None:
        export_dir = os.path.join(PROJECT_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = os.path.join(export_dir, 'precios_farmaciasvazquez.xlsx')

    from sqlalchemy import text
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = text("""
        SELECT * FROM (
            SELECT DISTINCT ON (p.id)
                p.nombre, p.url, p.categoria, p.ean,
                pr.precio, pr.precio_original, pr.en_stock, pr.fecha_captura
            FROM productos p
            JOIN precios pr ON p.id = pr.producto_id
            WHERE p.farmacia = :farmacia
            ORDER BY p.id, pr.fecha_captura DESC
        ) t
        ORDER BY categoria, nombre
    """)

    result = db.execute(query, {"farmacia": FARMACIA_NOMBRE})
    rows = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Precios FarmaciasVazquez"

    # Estilo Morado/Lila para FarmaciasVazquez
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid') # Morado
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_par = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid') # Lila claro
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    font_stock_si = Font(name='Calibri', color='2E7D32')
    font_stock_no = Font(name='Calibri', color='C62828')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    headers = ['Farmacia', 'Nombre', 'Categoría', 'EAN/SKU', 'Precio (€)',
               'Precio Original (€)', 'Descuento', 'En Stock', 'URL', 'Fecha']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        nombre, url, categoria, ean = row[0], row[1], row[2], row[3]
        precio = float(row[4]) if row[4] else 0
        precio_original = float(row[5]) if row[5] else None
        en_stock, fecha = row[6], row[7]

        descuento = ""
        if precio_original and precio_original > 0 and precio < precio_original:
            pct = ((precio_original - precio) / precio_original) * 100
            descuento = f"-{pct:.0f}%"

        fill = fill_par if row_idx % 2 == 0 else fill_impar
        data = [
            FARMACIA_NOMBRE, nombre, categoria or '', ean or '',
            precio, precio_original, descuento,
            'Sí' if en_stock else 'No', url,
            fecha.strftime('%Y-%m-%d %H:%M') if fecha else '',
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = align_left

        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=5).alignment = align_right
        if precio_original:
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=6).alignment = align_right
        ws.cell(row=row_idx, column=7).alignment = align_center
        ws.cell(row=row_idx, column=7).font = Font(name='Calibri', color='C62828', bold=True)

        stock_cell = ws.cell(row=row_idx, column=8)
        stock_cell.alignment = align_center
        stock_cell.font = font_stock_si if en_stock else font_stock_no
        ws.cell(row=row_idx, column=10).alignment = align_center

    column_widths = {1: 17, 2: 55, 3: 30, 4: 16, 5: 14, 6: 18, 7: 12, 8: 11, 9: 60, 10: 18}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    wb.save(filename)
    logger.info(f"Exportados {len(rows)} productos a '{filename}'")
    return len(rows)


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Scraper de FarmaciasVazquez (Doofinder API)")
    parser.add_argument("--limit", type=int, default=0, help="Limitar a N productos (0=todos)")
    parser.add_argument("--export", action="store_true", help="Solo exportar datos existentes a Excel")
    parser.add_argument("--output", type=str, default=None, help="Ruta del archivo de salida")
    args = parser.parse_args()

    engine = get_engine()
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    try:
        if args.export:
            logger.info("Exportando datos existentes a Excel...")
            exportar_a_excel(db, filename=args.output)
        else:
            logger.info("=" * 60)
            logger.info("SCRAPING DE FARMACIASVAZQUEZ VÍA DOOFINDER API")
            logger.info("=" * 60)
            ejecutar_scraping(db, limit=args.limit)

            logger.info("")
            logger.info("=" * 60)
            logger.info("EXPORTANDO A EXCEL")
            logger.info("=" * 60)
            exportar_a_excel(db, filename=args.output)
    finally:
        db.close()


if __name__ == "__main__":
    main()
