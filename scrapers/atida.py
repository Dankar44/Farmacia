"""
Scraper de Atida (Mifarma) - Usa la API de Algolia.

Atida usa Algolia como motor de búsqueda (igual que DosFarma).
Las peticiones van directamente a los servidores de Algolia, no a atida.com,
por lo que la protección anti-bot (403) no aplica.

Credenciales Algolia (descubiertas por ingeniería inversa):
  - App ID:    M8GRS7KXGP
  - Index:     atida_es_es_products
  - API Key:   Secured API Key (pública, usada por el navegador)

Si la API Key caduca, se puede renovar automáticamente con:
    python main.py atida --refresh-key

Uso (desde la raíz del proyecto vía main.py):
    python main.py atida                     # Todos los productos (~30k, ~5 min)
    python main.py atida --limit 100         # Solo 100 productos (test)
    python main.py atida --export            # Solo exportar datos a Excel
    python main.py atida --refresh-key       # Renovar API Key + scrapear
"""

import os
import sys
import re
import time
import random
import logging
import argparse
import requests
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from sqlalchemy.orm import sessionmaker

# Directorio raíz del proyecto
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
from db_models import get_engine, Producto, Precio, Base

# ============================================================
# LOGGING
# ============================================================
LOG_DIR = os.path.join(PROJECT_ROOT, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, 'scraper_atida.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ============================================================
# CONFIGURACIÓN DE ALGOLIA PARA ATIDA
# ============================================================
# Atida expone su Algolia con estas credenciales públicas.
# La API Key es una "Secured API Key" (base64) que el navegador
# del usuario necesita para hacer las búsquedas desde JavaScript.
#
# ¿Y SI CAMBIAN LA API KEY?
# ==========================
# Ejecutar con --refresh-key para extraerla automáticamente.
# ============================================================

ALGOLIA_APP_ID = "M8GRS7KXGP"
ALGOLIA_API_KEY = "ZDFkYzBhZTRhMTZkYTUzZWU3YTg4MGIxNGM3MmRiYTNjZGY0YWYwYzdhYTZlNjRiZjIyYTllMzA3MThlYjdmZnRhZ0ZpbHRlcnM9"
ALGOLIA_INDEX = "atida_es_es_products"
ALGOLIA_URL = f"https://{ALGOLIA_APP_ID}-dsn.algolia.net/1/indexes/*/queries"

FARMACIA_NOMBRE = "Atida"
BASE_URL = "https://www.atida.com"

# Delay entre peticiones a la API (segundos)
MIN_DELAY = 0.3
MAX_DELAY = 1.0

# Productos por página de Algolia (máximo 1000)
HITS_PER_PAGE = 1000

# Cuántos productos guardar por lote antes de commit
BATCH_SIZE = 100


# ============================================================
# EXTRACCIÓN AUTOMÁTICA DE LA API KEY (por si caduca)
# ============================================================
def extraer_api_key_fresca():
    """
    Visita atida.com con Playwright, busca la configuración de Algolia
    en el JavaScript de la página, y devuelve la API Key actual.
    """
    global ALGOLIA_API_KEY

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logger.error("Necesitas playwright instalado: pip install playwright && playwright install chromium")
        return None

    logger.info("Extrayendo API Key fresca de Atida...")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            viewport={"width": 1920, "height": 1080}
        )
        page = context.new_page()

        page.goto("https://www.atida.com/es-es/", wait_until="domcontentloaded", timeout=30000)
        time.sleep(5)

        result = page.evaluate("""() => {
            // Buscar en los scripts de la página
            const scripts = document.querySelectorAll('script');
            for (const script of scripts) {
                const text = script.textContent || '';
                // Buscar la API Key (base64 larga)
                const keyMatch = text.match(/apiKey['":\\s]+['"]([A-Za-z0-9=+/]{40,})['"]/i);
                const appMatch = text.match(/applicationId['":\\s]+['"]([A-Z0-9]+)['"]/i);
                if (keyMatch) {
                    return {
                        apiKey: keyMatch[1],
                        appId: appMatch ? appMatch[1] : null,
                    };
                }
            }
            
            // Buscar en window.__NEXT_DATA__
            if (window.__NEXT_DATA__) {
                const json = JSON.stringify(window.__NEXT_DATA__);
                const keyMatch = json.match(/"apiKey":"([^"]+)"/);
                const appMatch = json.match(/"applicationId":"([^"]+)"/);
                if (keyMatch) {
                    return {apiKey: keyMatch[1], appId: appMatch ? appMatch[1] : null};
                }
            }
            
            return null;
        }""")

        browser.close()

    if result and result.get("apiKey"):
        ALGOLIA_API_KEY = result["apiKey"]
        logger.info(f"API Key extraída correctamente: {ALGOLIA_API_KEY[:20]}...")
        logger.info(f"App ID confirmado: {result.get('appId', 'no encontrado')}")
        return ALGOLIA_API_KEY
    else:
        logger.error("No se pudo extraer la API Key. Usa la key manual.")
        return None


# ============================================================
# HEADERS PARA LA API DE ALGOLIA
# ============================================================
def get_algolia_headers():
    return {
        "x-algolia-application-id": ALGOLIA_APP_ID,
        "x-algolia-api-key": ALGOLIA_API_KEY,
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://www.atida.com/",
        "Origin": "https://www.atida.com",
    }


# ============================================================
# CONSULTA A ALGOLIA
# ============================================================
def consultar_algolia(query="", page=0, hits_per_page=HITS_PER_PAGE):
    """
    Hace una consulta a la API de Algolia de Atida.
    Si query="" busca TODOS los productos (browse).
    """
    payload = {
        "requests": [
            {
                "indexName": ALGOLIA_INDEX,
                "query": query,
                "params": f"hitsPerPage={hits_per_page}&page={page}"
            }
        ]
    }

    try:
        response = requests.post(
            ALGOLIA_URL,
            json=payload,
            headers=get_algolia_headers(),
            timeout=30
        )
    except requests.exceptions.RequestException as e:
        logger.error(f"Error de conexión con Algolia: {e}")
        return None

    if response.status_code != 200:
        logger.error(f"Error API Algolia: HTTP {response.status_code}")
        logger.error(response.text[:500])
        return None

    data = response.json()
    return data.get("results", [{}])[0]


# ============================================================
# EXTRACCIÓN DE DATOS DE ALGOLIA
# ============================================================
def extraer_datos_producto(hit):
    """Extrae datos relevantes de un 'hit' de Algolia de Atida."""
    nombre = hit.get("name", "Sin nombre")
    url = hit.get("url", "")
    sku = hit.get("sku", "")
    en_stock = bool(hit.get("in_stock", 0))
    brand = hit.get("brand", {})
    marca = brand if isinstance(brand, str) else brand.get("name", "") if isinstance(brand, dict) else ""

    # Precio
    precio = None
    precio_original = None

    price_data = hit.get("price", {})
    if isinstance(price_data, dict):
        # Atida puede tener price.EUR.default o price.sale / price.original
        eur_data = price_data.get("EUR", price_data)
        if isinstance(eur_data, dict):
            # Intentar varios formatos
            for key in ["default", "sale", "final", "value"]:
                val = eur_data.get(key)
                if val is not None:
                    try:
                        precio = Decimal(str(val))
                        break
                    except (InvalidOperation, ValueError):
                        continue

            # Precio original
            for key in ["default_original_formated", "original", "was", "default_formated"]:
                orig = eur_data.get(key)
                if orig is not None:
                    if isinstance(orig, str):
                        orig = orig.replace("€", "").replace(".", "").replace(",", ".").strip()
                    try:
                        precio_original = Decimal(str(orig))
                        if precio and precio_original <= precio:
                            precio_original = None  # No es descuento real
                        break
                    except (InvalidOperation, ValueError):
                        continue
        elif isinstance(eur_data, (int, float)):
            try:
                precio = Decimal(str(eur_data))
            except (InvalidOperation, ValueError):
                pass

    # Categorías
    categorias = hit.get("categories", {})
    categoria = ""
    if isinstance(categorias, dict):
        nivel0 = categorias.get("level0", [])
        if nivel0:
            categoria = nivel0[0] if isinstance(nivel0, list) else str(nivel0)
    elif isinstance(categorias, list) and categorias:
        categoria = categorias[0] if isinstance(categorias[0], str) else str(categorias[0])

    # URL completa
    if url and not url.startswith("http"):
        url = f"{BASE_URL}{url}"

    return {
        "nombre": nombre,
        "url": url,
        "sku": sku,
        "precio": precio,
        "precio_original": precio_original,
        "en_stock": en_stock,
        "categoria": categoria,
        "marca": marca,
    }


# ============================================================
# SCRAPING COMPLETO VÍA ALGOLIA
# ============================================================
def ejecutar_scraping(db, limit=0):
    """Extrae TODOS los productos de Atida vía la API de Algolia."""

    # Consultar cuántos productos hay
    resultado_inicial = consultar_algolia(query="", page=0, hits_per_page=1)
    if not resultado_inicial:
        logger.error("No se pudo conectar a la API de Algolia.")
        logger.error("Prueba con --refresh-key para renovar la API Key.")
        return

    total_productos = resultado_inicial.get("nbHits", 0)
    total_paginas = -(-total_productos // HITS_PER_PAGE)

    logger.info(f"Total de productos en Algolia: {total_productos}")
    logger.info(f"Total de páginas (de {HITS_PER_PAGE} productos): {total_paginas}")

    if limit > 0:
        paginas_a_procesar = min(total_paginas, -(-limit // HITS_PER_PAGE))
        logger.info(f"Limitando a {limit} productos ({paginas_a_procesar} páginas)")
    else:
        paginas_a_procesar = total_paginas

    exitos = 0
    errores = 0
    hoy = datetime.now(timezone.utc)

    for pagina in range(paginas_a_procesar):
        logger.info(f"Descargando página {pagina + 1}/{paginas_a_procesar}...")

        if pagina > 0:
            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        resultado = consultar_algolia(query="", page=pagina, hits_per_page=HITS_PER_PAGE)
        if not resultado:
            logger.error(f"Error en página {pagina + 1}. Reintentando...")
            time.sleep(5)
            resultado = consultar_algolia(query="", page=pagina, hits_per_page=HITS_PER_PAGE)
            if not resultado:
                logger.error(f"Fallo definitivo en página {pagina + 1}. Saltando.")
                continue

        hits = resultado.get("hits", [])
        logger.info(f"  Productos en esta página: {len(hits)}")

        if len(hits) == 0:
            logger.info("No hay más productos. Finalizando.")
            break

        for hit in hits:
            if limit > 0 and exitos >= limit:
                break

            datos = extraer_datos_producto(hit)

            if not datos["url"] or datos["precio"] is None:
                errores += 1
                continue

            # Buscar si ya existe
            producto = db.query(Producto).filter_by(url=datos["url"]).first()

            if not producto:
                producto = Producto(
                    nombre=datos["nombre"],
                    url=datos["url"],
                    farmacia=FARMACIA_NOMBRE,
                    categoria=datos["categoria"],
                    ean=datos["sku"],
                )
                db.add(producto)
                db.flush()
            else:
                producto.nombre = datos["nombre"]
                producto.categoria = datos["categoria"]
                if datos["sku"]:
                    producto.ean = datos["sku"]

            precio_record = Precio(
                producto_id=producto.id,
                precio=datos["precio"],
                precio_original=datos["precio_original"],
                en_stock=datos["en_stock"],
                fecha_captura=hoy,
            )
            db.add(precio_record)
            exitos += 1

            if exitos % BATCH_SIZE == 0:
                db.commit()
                logger.info(f"  Guardados: {exitos} productos")

        if limit > 0 and exitos >= limit:
            break

    db.commit()

    logger.info("=" * 60)
    logger.info("RESUMEN FINAL")
    logger.info("=" * 60)
    logger.info(f"Productos con precio extraído: {exitos}")
    logger.info(f"Productos sin precio/URL: {errores}")
    logger.info(f"Total en DB: {db.query(Producto).filter_by(farmacia=FARMACIA_NOMBRE).count()}")


# ============================================================
# EXPORTAR A EXCEL
# ============================================================
def exportar_a_excel(db, filename=None):
    if filename is None:
        export_dir = os.path.join(PROJECT_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = os.path.join(export_dir, 'precios_atida.xlsx')

    from sqlalchemy import text
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = text("""
        SELECT
            p.nombre, p.url, p.categoria, p.ean,
            pr.precio, pr.precio_original, pr.en_stock, pr.fecha_captura
        FROM productos p
        JOIN precios pr ON p.id = pr.producto_id
        WHERE p.farmacia = :farmacia
        AND pr.id = (
            SELECT pr2.id FROM precios pr2
            WHERE pr2.producto_id = p.id
            ORDER BY pr2.fecha_captura DESC
            LIMIT 1
        )
        ORDER BY p.categoria, p.nombre
    """)

    result = db.execute(query, {"farmacia": FARMACIA_NOMBRE})
    rows = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Precios Atida"

    # Estilo verde Atida
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='00796B', end_color='00796B', fill_type='solid')  # Teal/Verde Atida
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_par = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    font_stock_si = Font(name='Calibri', color='2E7D32')
    font_stock_no = Font(name='Calibri', color='C62828')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    headers = ['Farmacia', 'Nombre', 'Categoría', 'EAN/SKU', 'Precio (€)',
               'Precio Original (€)', 'Descuento', 'En Stock', 'URL', 'Fecha']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        nombre, url, categoria, ean = row[0], row[1], row[2], row[3]
        precio = float(row[4]) if row[4] else 0
        precio_original = float(row[5]) if row[5] else None
        en_stock, fecha = row[6], row[7]

        descuento = ""
        if precio_original and precio_original > 0 and precio < precio_original:
            pct = ((precio_original - precio) / precio_original) * 100
            descuento = f"-{pct:.0f}%"

        fill = fill_par if row_idx % 2 == 0 else fill_impar
        data = [
            FARMACIA_NOMBRE, nombre, categoria or '', ean or '',
            precio, precio_original, descuento,
            'Sí' if en_stock else 'No', url,
            fecha.strftime('%Y-%m-%d %H:%M') if fecha else '',
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = align_left

        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=5).alignment = align_right
        if precio_original:
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=6).alignment = align_right
        ws.cell(row=row_idx, column=7).alignment = align_center
        ws.cell(row=row_idx, column=7).font = Font(name='Calibri', color='C62828', bold=True)

        stock_cell = ws.cell(row=row_idx, column=8)
        stock_cell.alignment = align_center
        stock_cell.font = font_stock_si if en_stock else font_stock_no
        ws.cell(row=row_idx, column=10).alignment = align_center

    column_widths = {1: 12, 2: 55, 3: 30, 4: 16, 5: 14, 6: 18, 7: 12, 8: 11, 9: 60, 10: 18}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    wb.save(filename)
    logger.info(f"Exportados {len(rows)} productos a '{filename}'")
    return len(rows)


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Scraper de precios Atida (Algolia API)")
    parser.add_argument("--limit", type=int, default=0, help="Limitar a N productos (0=todos)")
    parser.add_argument("--export", action="store_true", help="Solo exportar datos existentes a Excel")
    parser.add_argument("--refresh-key", action="store_true", help="Renovar la API Key de Algolia")
    parser.add_argument("--output", type=str, default=None, help="Ruta del archivo de salida")
    args = parser.parse_args()

    engine = get_engine()
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    try:
        if args.refresh_key:
            logger.info("Renovando API Key de Algolia...")
            nueva_key = extraer_api_key_fresca()
            if not nueva_key:
                logger.error("No se pudo renovar la API Key.")
                return

        if args.export:
            logger.info("Exportando datos existentes a Excel...")
            exportar_a_excel(db, filename=args.output)
        else:
            logger.info("=" * 60)
            logger.info("SCRAPING DE ATIDA VÍA ALGOLIA API")
            logger.info("=" * 60)
            ejecutar_scraping(db, limit=args.limit)

            logger.info("")
            logger.info("=" * 60)
            logger.info("EXPORTANDO A EXCEL")
            logger.info("=" * 60)
            exportar_a_excel(db, filename=args.output)
    finally:
        db.close()


if __name__ == "__main__":
    main()
