"""
Scraper de DosFarma v3 - Usa la API de Algolia (descubierta por ingeniería inversa).

¡Este es el enfoque ganador! DosFarma usa Algolia como motor de búsqueda.
Podemos hacer peticiones directas a esta API y obtener precio, nombre, stock, etc.
de TODOS los productos sin tocar Cloudflare.

Estrategia:
1. Consultar la API de Algolia para obtener todos los productos paginados
2. Guardar en PostgreSQL (tabla productos + precios)
3. Exportar a Excel/CSV

Uso (desde la raíz del proyecto vía main.py):
    python main.py dosfarma                    # Scrapping completo
    python main.py dosfarma --limit 100        # Solo 100 productos (test)
    python main.py dosfarma --export           # Solo exportar a Excel
    python main.py dosfarma --refresh-key      # Renovar API Key + scraping
"""

import os
import sys
import requests
import time
import random
import logging
import argparse
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from sqlalchemy.orm import sessionmaker
from sqlalchemy import func, cast, Date

# Directorio raiz del proyecto (un nivel arriba de scrapers/)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
from db_models import get_engine, Producto, Precio, Base

# ============================================================
# LOGGING
# ============================================================
LOG_DIR = os.path.join(PROJECT_ROOT, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, 'scraper.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURACIÓN ALGOLIA (Descubierta por ingeniería inversa)
# ============================================================
#
# ¿CÓMO DESCUBRIMOS ESTO?
# ========================
# 1. Abrimos https://www.dosfarma.com/ en Chrome
# 2. Abrimos las DevTools (F12) → pestaña "Network"
# 3. Escribimos algo en la barra de búsqueda (ej: "gingilacer")
# 4. En Network, vimos peticiones POST a "algolia.net"
# 5. Clicamos en esa petición y vimos:
#    - URL: https://5FYR88UN93-dsn.algolia.net/1/indexes/*/queries
#    - Headers: x-algolia-application-id y x-algolia-api-key
#    - Body: {"requests": [{"indexName": "pro_dosfarma_es_products", ...}]}
#    - Response: JSON con nombre, precio, stock, URL de cada producto
#
# ¿POR QUÉ FUNCIONA?
# ===================
# Algolia es un servicio EXTERNO de búsqueda. Las peticiones van directamente
# a los servidores de Algolia (no a dosfarma.com), por lo que Cloudflare
# NO puede bloquearlas. La API Key es pública porque el navegador del usuario
# necesita usarla para hacer las búsquedas desde JavaScript.
#
# ¿Y SI CAMBIAN LA API KEY?
# ==========================
# La API Key que usamos es una "Secured API Key" que puede cambiar.
# Si deja de funcionar (error 403), hay que:
#   Opción A: Ejecutar con --refresh-key para extraerla automáticamente
#   Opción B: Ir a dosfarma.com → F12 → Network → buscar algo → copiar la key
#
# Lo que NO cambiará (casi nunca):
#   - ALGOLIA_APP_ID: Es el ID de la cuenta de DosFarma en Algolia
#   - ALGOLIA_INDEX: Es el nombre del índice de productos
#   - La estructura de la respuesta JSON
#
# ¿CÓMO REPLICAR ESTO EN OTRA WEB?
# ==================================
# 1. Abrir la web objetivo en Chrome con DevTools → Network
# 2. Buscar algo en la barra de búsqueda
# 3. Buscar peticiones a: algolia.net, doofinder.com, elasticsearch, /api/search
# 4. Si encuentras una, copiar: URL, headers, y body de la petición
# 5. Replicar esa petición en Python con requests.post()
# 6. ¡Listo! Sin Cloudflare, sin proxies, sin navegador
#
# SERVICIOS DE BÚSQUEDA COMUNES EN E-COMMERCE ESPAÑOL:
# - Algolia (algolia.net) → DosFarma, Decathlon, muchos Magento/Prestashop
# - Doofinder (doofinder.com) → Muy popular en España
# - Elasticsearch (/api/search, /elasticsearch) → Webs custom
# - Searchanise → Común en Shopify
# - Klevu → Común en Magento
# ============================================================

ALGOLIA_APP_ID = "5FYR88UN93"
ALGOLIA_API_KEY = "MDcyZWIyZjVlOTk0YzRjMDg2ZTBiNmUzZTcyNWE3YjZhMGZkOWQwYmQ0NzE0NDcwNTc4MWI2ZTFmMzBmMGRmMHRhZ0ZpbHRlcnM9"
ALGOLIA_INDEX = "pro_dosfarma_es_products"
ALGOLIA_URL = f"https://{ALGOLIA_APP_ID}-dsn.algolia.net/1/indexes/*/queries"

FARMACIA_NOMBRE = "DosFarma"

# Delay entre peticiones a la API (segundos)
MIN_DELAY = 0.5
MAX_DELAY = 1.5

# Productos por página de Algolia (máximo 1000)
HITS_PER_PAGE = 1000

# Cuántos productos guardar por lote antes de commit
BATCH_SIZE = 100


# ============================================================
# EXTRACCIÓN AUTOMÁTICA DE LA API KEY (por si caduca)
# ============================================================
# Esta función abre DosFarma en un navegador real (Playwright),
# lee el JavaScript de la página, y extrae la API Key fresca.
# Se usa con: python scraper_dosfarma.py --refresh-key
# ============================================================
def extraer_api_key_fresca():
    """
    Visita dosfarma.com con Playwright, busca la configuración de Algolia
    en el JavaScript de la página, y devuelve la API Key actual.
    
    Cómo funciona:
    1. Abre un Chrome real (con stealth para pasar Cloudflare)
    2. Va a dosfarma.com
    3. Busca en los <script> de la página el objeto 'algoliaConfig'
    4. Extrae el campo 'apiKey' de ese objeto
    5. Devuelve la key para usarla en las peticiones
    """
    global ALGOLIA_API_KEY
    
    try:
        from playwright.sync_api import sync_playwright
        from playwright_stealth import Stealth
    except ImportError:
        logger.error("Necesitas playwright y playwright-stealth instalados.")
        logger.error("pip install playwright playwright-stealth && playwright install chromium")
        return None
    
    logger.info("Extrayendo API Key fresca de DosFarma...")
    
    with Stealth().use_sync(sync_playwright()) as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        
        # Visitar la página principal
        page.goto("https://www.dosfarma.com/", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(5000)  # Esperar a que cargue todo el JS
        
        # Extraer la configuración de Algolia del JavaScript de la página
        # DosFarma la guarda en window.algoliaConfig
        result = page.evaluate("""() => {
            // Método 1: Buscar en window.algoliaConfig (lo más fiable)
            if (window.algoliaConfig) {
                return {
                    apiKey: window.algoliaConfig.apiKey || null,
                    appId: window.algoliaConfig.applicationId || null,
                };
            }
            
            // Método 2: Buscar en los scripts de la página
            const scripts = document.querySelectorAll('script');
            for (const script of scripts) {
                const text = script.textContent || '';
                const keyMatch = text.match(/apiKey['":\\s]*['"]([A-Za-z0-9=+/]+)['"]/);
                const appMatch = text.match(/applicationId['":\\s]*['"]([A-Z0-9]+)['"]/);
                if (keyMatch) {
                    return {
                        apiKey: keyMatch[1],
                        appId: appMatch ? appMatch[1] : null,
                    };
                }
            }
            
            return null;
        }""")
        
        browser.close()
    
    if result and result.get("apiKey"):
        ALGOLIA_API_KEY = result["apiKey"]
        logger.info(f"API Key extraída correctamente: {ALGOLIA_API_KEY[:20]}...")
        logger.info(f"App ID confirmado: {result.get('appId', 'no encontrado')}")
        return ALGOLIA_API_KEY
    else:
        logger.error("No se pudo extraer la API Key. Usa la key manual.")
        return None


# ============================================================
# HEADERS PARA LA API DE ALGOLIA
# ============================================================
def get_algolia_headers():
    return {
        "x-algolia-application-id": ALGOLIA_APP_ID,
        "x-algolia-api-key": ALGOLIA_API_KEY,
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": "https://www.dosfarma.com/",
        "Origin": "https://www.dosfarma.com",
    }


# ============================================================
# CONSULTA A ALGOLIA
# ============================================================
def consultar_algolia(query="", page=0, hits_per_page=HITS_PER_PAGE):
    """
    Hace una consulta a la API de Algolia de DosFarma.
    Si query="" busca TODOS los productos (browse).
    """
    payload = {
        "requests": [
            {
                "indexName": ALGOLIA_INDEX,
                "query": query,
                "params": f"hitsPerPage={hits_per_page}&page={page}&numericFilters=visibility_search%3D1"
            }
        ]
    }
    
    response = requests.post(
        ALGOLIA_URL,
        json=payload,
        headers=get_algolia_headers(),
        timeout=30
    )
    
    if response.status_code != 200:
        logger.error(f"Error API Algolia: HTTP {response.status_code}")
        logger.error(response.text[:500])
        return None
    
    data = response.json()
    return data.get("results", [{}])[0]


# ============================================================
# EXTRACCIÓN DE DATOS DE ALGOLIA
# ============================================================
def extraer_datos_producto(hit):
    """Extrae datos relevantes de un 'hit' de Algolia."""
    nombre = hit.get("name", "Sin nombre")
    url = hit.get("url", "")
    sku = hit.get("sku", "")
    en_stock = bool(hit.get("in_stock", 0))
    
    # Precio
    precio = None
    precio_original = None
    
    price_data = hit.get("price", {}).get("EUR", {})
    if price_data:
        try:
            precio = Decimal(str(price_data.get("default", 0)))
        except (InvalidOperation, ValueError):
            pass
        
        # Precio original (antes del descuento)
        original_str = price_data.get("default_original_formated", "")
        if original_str:
            # Formato: "14,93 €" -> 14.93
            original_str = original_str.replace("€", "").replace(".", "").replace(",", ".").strip()
            try:
                precio_original = Decimal(original_str)
            except (InvalidOperation, ValueError):
                pass
    
    # Categorías
    categorias = hit.get("categories", {})
    categoria = ""
    if isinstance(categorias, dict):
        # Tomar la primera categoría de nivel 0
        nivel0 = categorias.get("level0", [])
        if nivel0:
            categoria = nivel0[0] if isinstance(nivel0, list) else str(nivel0)
    
    return {
        "nombre": nombre,
        "url": f"https://www.dosfarma.com{url}" if url and not url.startswith("http") else url,
        "sku": sku,
        "precio": precio,
        "precio_original": precio_original,
        "en_stock": en_stock,
        "categoria": categoria,
    }


# ============================================================
# SCRAPING COMPLETO VÍA ALGOLIA
# ============================================================
def ejecutar_scraping_algolia(db, limit=0):
    """Extrae TODOS los productos de DosFarma vía la API de Algolia."""
    
    # Primero, consultar cuántos productos hay
    resultado_inicial = consultar_algolia(query="", page=0, hits_per_page=1)
    if not resultado_inicial:
        logger.error("No se pudo conectar a la API de Algolia.")
        return
    
    total_productos = resultado_inicial.get("nbHits", 0)
    # Calcular páginas correctamente: total_productos / productos_por_página
    total_paginas = -(-total_productos // HITS_PER_PAGE)  # ceil division
    
    logger.info(f"Total de productos en Algolia: {total_productos}")
    logger.info(f"Total de páginas (de {HITS_PER_PAGE} productos): {total_paginas}")
    
    if limit > 0:
        paginas_a_procesar = min(total_paginas, -(-limit // HITS_PER_PAGE))
        logger.info(f"Limitando a {limit} productos ({paginas_a_procesar} páginas)")
    else:
        paginas_a_procesar = total_paginas
    
    exitos = 0
    errores = 0
    hoy = datetime.now(timezone.utc)
    
    for pagina in range(paginas_a_procesar):
        logger.info(f"Descargando página {pagina + 1}/{paginas_a_procesar}...")
        
        # Delay entre peticiones
        if pagina > 0:
            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))
        
        resultado = consultar_algolia(query="", page=pagina, hits_per_page=HITS_PER_PAGE)
        if not resultado:
            logger.error(f"Error en página {pagina + 1}. Reintentando...")
            time.sleep(5)
            resultado = consultar_algolia(query="", page=pagina, hits_per_page=HITS_PER_PAGE)
            if not resultado:
                logger.error(f"Fallo definitivo en página {pagina + 1}. Saltando.")
                continue
        
        hits = resultado.get("hits", [])
        logger.info(f"  Productos en esta página: {len(hits)}")
        
        # Si no hay más resultados, parar
        if len(hits) == 0:
            logger.info("No hay más productos. Finalizando.")
            break
        
        for hit in hits:
            if limit > 0 and exitos >= limit:
                break
            
            datos = extraer_datos_producto(hit)
            
            if not datos["url"] or datos["precio"] is None:
                errores += 1
                continue
            
            # Buscar si el producto ya existe en la DB
            producto = db.query(Producto).filter_by(url=datos["url"]).first()
            
            if not producto:
                producto = Producto(
                    nombre=datos["nombre"],
                    url=datos["url"],
                    farmacia=FARMACIA_NOMBRE,
                    categoria=datos["categoria"],
                    ean=datos["sku"],
                )
                db.add(producto)
                db.flush()  # Para obtener el ID generado
            else:
                # Actualizar nombre y categoría si han cambiado
                producto.nombre = datos["nombre"]
                producto.categoria = datos["categoria"]
                if datos["sku"]:
                    producto.ean = datos["sku"]
            
            # Guardar precio
            precio_record = Precio(
                producto_id=producto.id,
                precio=datos["precio"],
                precio_original=datos["precio_original"],
                en_stock=datos["en_stock"],
                fecha_captura=hoy,
            )
            db.add(precio_record)
            exitos += 1
            
            # Commit por lotes
            if exitos % BATCH_SIZE == 0:
                db.commit()
                logger.info(f"  Guardados: {exitos} productos")
        
        if limit > 0 and exitos >= limit:
            break
    
    # Commit final
    db.commit()
    
    # Resumen
    logger.info("=" * 60)
    logger.info("RESUMEN FINAL")
    logger.info("=" * 60)
    logger.info(f"Productos con precio extraído: {exitos}")
    logger.info(f"Productos sin precio/URL: {errores}")
    logger.info(f"Total en DB: {db.query(Producto).filter_by(farmacia=FARMACIA_NOMBRE).count()}")


# ============================================================
# EXPORTAR A EXCEL (.xlsx) CON FORMATO BONITO
# ============================================================
def exportar_a_excel(db, filename=None):
    if filename is None:
        export_dir = os.path.join(PROJECT_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = os.path.join(export_dir, 'precios_dosfarma.xlsx')
    """Exporta todos los precios más recientes a un Excel con formato profesional."""
    from sqlalchemy import text
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    
    # Query para obtener el precio más reciente de cada producto
    query = text("""
        SELECT 
            p.nombre, 
            p.url, 
            p.categoria, 
            p.ean,
            pr.precio, 
            pr.precio_original, 
            pr.en_stock,
            pr.fecha_captura
        FROM productos p
        JOIN precios pr ON p.id = pr.producto_id
        WHERE p.farmacia = :farmacia
        AND pr.id = (
            SELECT pr2.id FROM precios pr2 
            WHERE pr2.producto_id = p.id 
            ORDER BY pr2.fecha_captura DESC 
            LIMIT 1
        )
        ORDER BY p.categoria, p.nombre
    """)
    
    result = db.execute(query, {"farmacia": FARMACIA_NOMBRE})
    rows = result.fetchall()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios DosFarma"
    
    # ---- ESTILOS ----
    # Cabecera
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Filas alternas (zebra)
    fill_par = PatternFill(start_color='F0F7FA', end_color='F0F7FA', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Stock
    fill_stock_si = Font(name='Calibri', color='2E7D32')  # Verde
    fill_stock_no = Font(name='Calibri', color='C62828')  # Rojo
    
    # Bordes finos
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    
    # Alineaciones
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=False)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # ---- CABECERA ----
    headers = ['Farmacia', 'Nombre', 'Categoría', 'EAN/SKU', 'Precio (€)', 'Precio Original (€)', 'Descuento', 'En Stock', 'URL', 'Fecha']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # ---- DATOS ----
    for row_idx, row in enumerate(rows, 2):
        nombre = row[0]
        url = row[1]
        categoria = row[2]
        ean = row[3]
        precio = float(row[4]) if row[4] else 0
        precio_original = float(row[5]) if row[5] else None
        en_stock = row[6]
        fecha = row[7]
        
        # Calcular descuento
        descuento = ""
        if precio_original and precio_original > 0 and precio < precio_original:
            pct = ((precio_original - precio) / precio_original) * 100
            descuento = f"-{pct:.0f}%"
        
        # Escribir datos
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        
        data = [
            FARMACIA_NOMBRE,
            nombre,
            categoria,
            ean or '',
            precio,
            precio_original,
            descuento,
            'Sí' if en_stock else 'No',
            url,
            fecha.strftime('%Y-%m-%d %H:%M') if fecha else '',
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = align_left
        
        # Formato especial por columna
        ws.cell(row=row_idx, column=1).alignment = align_center  # Farmacia
        ws.cell(row=row_idx, column=4).alignment = align_center  # EAN
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00 €'  # Precio
        ws.cell(row=row_idx, column=5).alignment = align_right
        if precio_original:
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00 €'  # Precio original
        ws.cell(row=row_idx, column=6).alignment = align_right
        ws.cell(row=row_idx, column=7).alignment = align_center  # Descuento
        ws.cell(row=row_idx, column=7).font = Font(name='Calibri', color='C62828', bold=True)  # Descuento en rojo
        
        # Stock con color
        stock_cell = ws.cell(row=row_idx, column=8)
        stock_cell.alignment = align_center
        stock_cell.font = fill_stock_si if en_stock else fill_stock_no
        
        ws.cell(row=row_idx, column=10).alignment = align_center  # Fecha
    
    # ---- ANCHO DE COLUMNAS (auto-ajuste) ----
    column_widths = {
        1: 12,   # Farmacia
        2: 55,   # Nombre (el más ancho)
        3: 25,   # Categoría
        4: 16,   # EAN/SKU
        5: 14,   # Precio
        6: 18,   # Precio Original
        7: 12,   # Descuento
        8: 11,   # En Stock
        9: 60,   # URL
        10: 18,  # Fecha
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    # ---- CONGELAR CABECERA (para que siempre se vea al hacer scroll) ----
    ws.freeze_panes = 'A2'
    
    # ---- AUTOFILTRO (para poder filtrar por categoría, stock, etc.) ----
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"
    
    # Guardar
    wb.save(filename)
    logger.info(f"Exportados {len(rows)} productos a '{filename}'")
    return len(rows)


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Scraper de precios DosFarma (vía Algolia API)")
    parser.add_argument("--limit", type=int, default=0, help="Limitar a N productos (0=todos)")
    parser.add_argument("--export", action="store_true", help="Solo exportar datos existentes a Excel")
    parser.add_argument("--output", type=str, default=None, help="Ruta del archivo de salida (default: exports/precios_dosfarma.xlsx)")
    parser.add_argument("--refresh-key", action="store_true", help="Extraer API Key fresca de DosFarma (usar si da error 403)")
    args = parser.parse_args()
    
    # Si piden refrescar la API Key, hacerlo antes de nada
    if args.refresh_key:
        key = extraer_api_key_fresca()
        if not key:
            logger.error("No se pudo extraer la API Key. Abortando.")
            return
        logger.info("API Key actualizada. Continuando con el scraping...")
    
    engine = get_engine()
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()
    
    try:
        if args.export:
            # Solo exportar
            logger.info("Exportando datos existentes a Excel...")
            exportar_a_excel(db, filename=args.output)
        else:
            # Scraping + exportación
            logger.info("=" * 60)
            logger.info("SCRAPING DE DOSFARMA VÍA ALGOLIA API")
            logger.info("=" * 60)
            ejecutar_scraping_algolia(db, limit=args.limit)
            
            # Exportar automáticamente
            logger.info("")
            logger.info("=" * 60)
            logger.info("EXPORTANDO A EXCEL")
            logger.info("=" * 60)
            exportar_a_excel(db, filename=args.output)
    
    finally:
        db.close()


if __name__ == "__main__":
    main()

