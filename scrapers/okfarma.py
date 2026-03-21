"""
Scraper de Okfarma - Usa Playwright + Sitemap + Batch Fetch.

Okfarma tiene un catálogo público accesible vía sitemap y sin
protección anti-bot estricta.

Estrategia:
  1. Abrir un navegador real con Playwright UNA SOLA VEZ
  2. Descargar el sitemap (`sitemap-1.xml`) y extraer sitemaps de productos
  3. Desde el navegador, hacer fetch() al sitemap XML → obtener todas las URLs
  4. Batch-fetch de páginas de producto con Promise.all() (20 a la vez)
  5. Parsear precios del HTML con expresiones regulares (sin renderizar DOM)
  6. Guardar en PostgreSQL y exportar a Excel

Uso (desde la raíz del proyecto vía main.py):
    python main.py okfarma                  # Scraping completo
    python main.py okfarma --limit 1000     # Solo 1000 productos (test)
    python main.py okfarma --export         # Solo exportar a Excel
"""

import os
import sys
import re
import time
import logging
import argparse
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from sqlalchemy.orm import sessionmaker

# Directorio raíz del proyecto
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
from db_models import get_engine, Producto, Precio, Base

# ============================================================
# LOGGING
# ============================================================
LOG_DIR = os.path.join(PROJECT_ROOT, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, 'scraper_okfarma.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ============================================================
# CONFIGURACIÓN
# ============================================================
FARMACIA_NOMBRE = "Okfarma"
BASE_URL = "https://okfarma.es"
SITEMAP_INDEX = "https://okfarma.es/sitemap-1.xml"
BATCH_SIZE_FETCH = 20   # Productos por batch de fetch (Promise.all)
BATCH_SIZE_DB = 200     # Productos por commit a la DB
DELAY_BETWEEN_BATCHES = 0.3  # Segundos entre batches (para no saturar)

# JavaScript para extraer datos rapidísimo (ejecutado en entorno browser)
EXTRACT_FROM_HTML_JS = r"""
(urls) => {
    // Helper: decodificar entidades HTML
    function decodeEntities(s) {
        if (!s) return s;
        return s.replace(/&#x([0-9a-f]+);/gi, (_, hex) => String.fromCharCode(parseInt(hex, 16)))
                .replace(/&#(\d+);/g, (_, dec) => String.fromCharCode(parseInt(dec)))
                .replace(/&amp;/g, '&').replace(/&lt;/g, '<').replace(/&gt;/g, '>')
                .replace(/&quot;/g, '"').replace(/&apos;/g, "'");
    }

    return Promise.all(urls.map(url =>
        fetch(url)
            .then(r => r.text())
            .then(html => {
                const data = {url: url, ok: true};

                // 1. Nombre
                const titleMatch = html.match(/<h1[^>]*>([^<]+)<\/h1>/i);
                if (titleMatch) {
                    data.nombre = decodeEntities(titleMatch[1].trim());
                } else {
                    const tMatch = html.match(/<title>([^<]+)<\/title>/i);
                    if (tMatch) data.nombre = decodeEntities(tMatch[1].split('|')[0].trim());
                }

                // 2. Precio
                let precio = null;
                const priceMatch1 = html.match(/product:price:amount"\s*content="([\d.]+)"/i);
                if (priceMatch1) {
                    precio = parseFloat(priceMatch1[1]);
                } else {
                    const priceMatch2 = html.match(/price_pvp[^>]*content="([\d.]+)"/i);
                    if (priceMatch2) {
                        precio = parseFloat(priceMatch2[1]);
                    } else {
                        const jsonPrice = html.match(/"price"\s*:\s*"?([\d.]+)"?/i);
                        if (jsonPrice) precio = parseFloat(jsonPrice[1]);
                    }
                }
                data.precio = precio;

                // 3. EAN
                let sku = "";
                const eanMatch1 = html.match(/"ean13"\s*:\s*"(\d{13})"/);
                if (eanMatch1) {
                    sku = eanMatch1[1];
                } else {
                    const eanMatch2 = html.match(/ean13[^\d]*(\d{13})/i);
                    if (eanMatch2) {
                        sku = eanMatch2[1];
                    } else {
                        const referenceMatch = html.match(/Referencia.*?(\d{5,8})/i);
                        if (referenceMatch) {
                            sku = referenceMatch[1].trim();
                        } else {
                            const fallbackSku = html.match(/<span[^>]*itemprop="sku"[^>]*>([^<]+)<\/span>/i);
                            if (fallbackSku) sku = fallbackSku[1].trim();
                        }
                    }
                }
                data.sku = sku;

                // 4. Stock
                data.enStock = !html.includes('Agotado') && !html.includes('Fuera de stock');

                return data;
            })
            .catch(e => ({url: url, ok: false, error: e.toString()}))
    ));
}
"""


# ============================================================
# FASE 1: DESCUBRIR PRODUCTOS VÍA SITEMAP
# ============================================================
def descubrir_productos_sitemap(page):
    """Descarga el sitemap index y obtiene las URLs de todos los productos."""
    logger.info("Descargando sitemap index...")
    sitemap_xml = page.evaluate("""(url) => {
        return fetch(url).then(r => r.text()).catch(e => "Error: " + e.toString());
    }""", SITEMAP_INDEX)

    if sitemap_xml.startswith("Error"):
        logger.error(f"Error descargando sitemap: {sitemap_xml}")
        return {}

    sub_urls = re.findall(r'<loc>(https://[^<]+)</loc>', sitemap_xml)
    product_subs = [u for u in sub_urls if 'sitemap-products-' in u]
    logger.info(f"Encontrados {len(product_subs)} sub-sitemaps de productos")

    productos_por_categoria = {}
    total = 0

    for i, sub_url in enumerate(product_subs):
        logger.info(f"  Procesando sub-sitemap {sub_url}...")
        sub_xml = page.evaluate("""(url) => {
            return fetch(url).then(r => r.text()).catch(e => "Error: " + e.toString());
        }""", sub_url)

        if sub_xml.startswith("Error"):
            logger.warning(f"  Error en sub-sitemap {sub_url}: {sub_xml}")
            continue

        urls = re.findall(r'<loc>(https://[^<]+)</loc>', sub_xml)
        
        valid_urls = []
        for u in urls:
            if not u.endswith('.jpg') and not u.endswith('.png'):
                valid_urls.append(u)
        
        if valid_urls:
            # Okfarma agrupa por número en sitemap, ponemos "General" de cat.
            cat_name = f"Productos {i+1}"
            productos_por_categoria[cat_name] = valid_urls
            total += len(valid_urls)
            logger.info(f"  [{i+1}/{len(product_subs)}] -> {len(valid_urls)} productos (acumulado: {total})")

    logger.info(f"Total URLs descubiertas: {total}")
    return productos_por_categoria


# ============================================================
# FASE 2: EXTRAER PRECIOS EN LOTES
# ============================================================
def extraer_precios_batch(page, db, productos_por_categoria, limit=0):
    exitos = 0
    errores = 0
    sin_precio = 0
    hoy = datetime.now(timezone.utc)

    # Cargar URLs ya procesadas
    productos_existentes = db.query(Producto.url).filter_by(farmacia=FARMACIA_NOMBRE).all()
    urls_existentes = {p[0] for p in productos_existentes}
    logger.info(f"Productos ya en DB: {len(urls_existentes)}")

    # Construir lista de URLs
    urls_pendientes = []
    urls_vistas = set()
    for cat_nombre, urls in productos_por_categoria.items():
        for url in urls:
            if url not in urls_existentes and url not in urls_vistas:
                urls_pendientes.append((url, cat_nombre))
                urls_vistas.add(url)

    total_pendientes = len(urls_pendientes)
    if limit > 0:
        urls_pendientes = urls_pendientes[:limit]
    logger.info(f"URLs pendientes de procesar: {len(urls_pendientes)} (de {total_pendientes} nuevas)")

    total_batches = -(-len(urls_pendientes) // BATCH_SIZE_FETCH)
    start_time = time.time()

    for batch_idx in range(total_batches):
        batch_start = batch_idx * BATCH_SIZE_FETCH
        batch_end = min(batch_start + BATCH_SIZE_FETCH, len(urls_pendientes))
        batch = urls_pendientes[batch_start:batch_end]
        batch_urls = [item[0] for item in batch]
        batch_cats = {item[0]: item[1] for item in batch}

        try:
            results = page.evaluate(EXTRACT_FROM_HTML_JS, batch_urls)
        except Exception as e:
            logger.error(f"Error en batch {batch_idx}: {e}")
            errores += len(batch_urls)
            time.sleep(2)
            continue

        for prod_data in results:
            prod_url = prod_data.get('url', '')

            if not prod_data.get('ok', False):
                errores += 1
                continue

            nombre = prod_data.get('nombre', '')
            precio_val = prod_data.get('precio')

            if precio_val is None:
                sin_precio += 1
                continue

            try:
                precio = Decimal(str(precio_val))
            except (InvalidOperation, ValueError):
                sin_precio += 1
                continue

            cat_nombre = batch_cats.get(prod_url, 'General')
            sku = prod_data.get('sku', '')
            en_stock = prod_data.get('enStock', True)

            producto_db = Producto(
                nombre=nombre,
                url=prod_url,
                farmacia=FARMACIA_NOMBRE,
                categoria=cat_nombre,
                ean=sku,
            )
            db.add(producto_db)
            db.flush()

            precio_record = Precio(
                producto_id=producto_db.id,
                precio=precio,
                precio_original=None,  # Okfarma scrape price is direct
                en_stock=en_stock,
                fecha_captura=hoy,
            )
            db.add(precio_record)
            exitos += 1

        if exitos > 0 and exitos % BATCH_SIZE_DB == 0:
            db.commit()

        if (batch_idx + 1) % 50 == 0 or batch_idx == total_batches - 1:
            elapsed = time.time() - start_time
            rate = exitos / elapsed if elapsed > 0 else 0
            eta = (len(urls_pendientes) - batch_end) / rate / 60 if rate > 0 else 0
            logger.info(
                f"  Progreso: {batch_end}/{len(urls_pendientes)} | "
                f"Éxitos: {exitos} | Errores: {errores} | Sin precio: {sin_precio} | "
                f"Velocidad: {rate:.1f} prod/s | ETA: {eta:.0f} min"
            )

        if DELAY_BETWEEN_BATCHES > 0:
            time.sleep(DELAY_BETWEEN_BATCHES)

    db.commit()
    return exitos, errores, sin_precio


# ============================================================
# EJECUTAR SCRAPING COMPLETO
# ============================================================
def ejecutar_scraping(db, limit=0):
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1920, "height": 1080}
        )
        page = context.new_page()

        logger.info("Abriendo página principal de Okfarma para iniciar contexto...")
        try:
            page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30000)
            time.sleep(2)
        except Exception as e:
            logger.error(f"Error cargando Okfarma: {e}")
            browser.close()
            return

        logger.info("\n" + "=" * 60)
        logger.info("FASE 1: DESCUBRIMIENTO VÍA SITEMAP")
        logger.info("=" * 60)
        productos_por_categoria = descubrir_productos_sitemap(page)

        total_urls = sum(len(v) for v in productos_por_categoria.values())
        if total_urls == 0:
            logger.error("No se encontraron productos en el sitemap.")
            browser.close()
            return

        logger.info("\n" + "=" * 60)
        logger.info("FASE 2: EXTRACCIÓN DE PRECIOS (BATCH FETCH)")
        logger.info("=" * 60)
        exitos, errores, sin_precio = extraer_precios_batch(page, db, productos_por_categoria, limit)

        browser.close()

    logger.info("\n" + "=" * 60)
    logger.info("RESUMEN FINAL")
    logger.info("=" * 60)
    logger.info(f"Productos extraídos con éxito: {exitos}")
    logger.info(f"Errores de fetch: {errores}")
    logger.info(f"Sin precio encontrado: {sin_precio}")
    logger.info(f"Total en DB: {db.query(Producto).filter_by(farmacia=FARMACIA_NOMBRE).count()}")


# ============================================================
# EXPORTAR A EXCEL
# ============================================================
def exportar_a_excel(db, filename=None):
    if filename is None:
        export_dir = os.path.join(PROJECT_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = os.path.join(export_dir, 'precios_okfarma.xlsx')

    from sqlalchemy import text
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = text("""
        SELECT
            p.nombre, p.url, p.categoria, p.ean,
            pr.precio, pr.precio_original, pr.en_stock, pr.fecha_captura
        FROM productos p
        JOIN precios pr ON p.id = pr.producto_id
        WHERE p.farmacia = :farmacia
        AND pr.id = (
            SELECT pr2.id FROM precios pr2
            WHERE pr2.producto_id = p.id
            ORDER BY pr2.fecha_captura DESC
            LIMIT 1
        )
        ORDER BY p.nombre
    """)

    result = db.execute(query, {"farmacia": FARMACIA_NOMBRE})
    rows = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Precios Okfarma"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='00bcd4', end_color='00bcd4', fill_type='solid') # Cyan okfarma
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_par = PatternFill(start_color='e0f7fa', end_color='e0f7fa', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    font_stock_si = Font(name='Calibri', color='2E7D32')
    font_stock_no = Font(name='Calibri', color='C62828')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    headers = ['Farmacia', 'Nombre', 'Categoría', 'ID Producto', 'Precio (€)',
               'Precio Original (€)', 'Descuento', 'En Stock', 'URL', 'Fecha']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        nombre, url, categoria, ean = row[0], row[1], row[2], row[3]
        precio = float(row[4]) if row[4] else 0
        precio_original = float(row[5]) if row[5] else None
        en_stock, fecha = row[6], row[7]

        descuento = ""
        if precio_original and precio_original > 0 and precio < precio_original:
            pct = ((precio_original - precio) / precio_original) * 100
            descuento = f"-{pct:.0f}%"

        fill = fill_par if row_idx % 2 == 0 else fill_impar
        data = [
            FARMACIA_NOMBRE, nombre, categoria or '', ean or '',
            precio, precio_original, descuento,
            'Sí' if en_stock else 'No', url,
            fecha.strftime('%Y-%m-%d %H:%M') if fecha else '',
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = align_left

        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=5).alignment = align_right
        if precio_original:
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00 €'
            ws.cell(row=row_idx, column=6).alignment = align_right
        ws.cell(row=row_idx, column=7).alignment = align_center
        ws.cell(row=row_idx, column=7).font = Font(name='Calibri', color='C62828', bold=True)

        stock_cell = ws.cell(row=row_idx, column=8)
        stock_cell.alignment = align_center
        stock_cell.font = font_stock_si if en_stock else font_stock_no
        ws.cell(row=row_idx, column=10).alignment = align_center

    column_widths = {1: 14, 2: 55, 3: 30, 4: 14, 5: 14, 6: 18, 7: 12, 8: 11, 9: 60, 10: 18}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    wb.save(filename)
    logger.info(f"Exportados {len(rows)} productos a '{filename}'")
    return len(rows)


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Scraper de precios Okfarma (Playwright + Sitemap)")
    parser.add_argument("--limit", type=int, default=0, help="Limitar a N productos (0=todos)")
    parser.add_argument("--export", action="store_true", help="Solo exportar datos existentes a Excel")
    parser.add_argument("--output", type=str, default=None, help="Ruta del archivo de salida")
    args = parser.parse_args()

    engine = get_engine()
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    try:
        if args.export:
            logger.info("Exportando datos existentes a Excel...")
            exportar_a_excel(db, filename=args.output)
        else:
            logger.info("=" * 60)
            logger.info("SCRAPING DE OKFARMA")
            logger.info("=" * 60)
            ejecutar_scraping(db, limit=args.limit)

            logger.info("")
            logger.info("=" * 60)
            logger.info("EXPORTANDO A EXCEL")
            logger.info("=" * 60)
            exportar_a_excel(db, filename=args.output)
    finally:
        db.close()


if __name__ == "__main__":
    main()
