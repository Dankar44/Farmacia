"""
Scraper de FarmaciasDirect - Usa la API de Empathy.co (descubierta por ingeniería inversa).

FarmaciasDirect usa Empathy.co como motor de búsqueda externo.
La API es accesible directamente sin protección anti-bot.

Estrategia:
  1. Descargar el sitemap XML para obtener la lista completa de productos
  2. Extraer el nombre de cada producto de la URL del sitemap
  3. Buscar grupos de productos en la API de Empathy.co
  4. Guardar en PostgreSQL y exportar a Excel

Uso (desde la raíz del proyecto vía main.py):
    python main.py farmaciasdirect                    # Scraping completo
    python main.py farmaciasdirect --limit 100        # Solo 100 productos (test)
    python main.py farmaciasdirect --export           # Solo exportar a Excel

¿CÓMO DESCUBRIMOS LA API?
==========================
1. Abrimos https://www.farmaciasdirect.com/ en Chrome
2. DevTools (F12) → Network
3. Escribimos algo en el buscador
4. Vimos peticiones GET a api.empathy.co
5. Endpoint: https://api.empathy.co/search/v1/query/farmaciasdirect/search
6. Parámetros: query, rows, start, lang, store, instance, scope
7. Respuesta JSON con: nombre, precioFinal, precioBase, disponibilidad, link, marca

NOTA: A diferencia de DosFarma (Algolia), esta API no permite "browse all".
Se necesita un término de búsqueda. Usamos el sitemap para extraer nombres
de productos y luego buscamos cada uno en la API.
"""

import os
import sys
import requests
import time
import random
import logging
import argparse
import re
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from xml.etree import ElementTree
from sqlalchemy.orm import sessionmaker
from sqlalchemy import func

# Directorio raíz del proyecto
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
from db_models import get_engine, Producto, Precio, Base

# ============================================================
# LOGGING
# ============================================================
LOG_DIR = os.path.join(PROJECT_ROOT, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, 'scraper_farmaciasdirect.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURACIÓN API EMPATHY.CO
# ============================================================
EMPATHY_BASE_URL = "https://api.empathy.co/search/v1/query/farmaciasdirect/search"
FARMACIA_NOMBRE = "FarmaciasDirect"

# Sitemap para descubrir todos los productos
SITEMAP_INDEX_URL = "https://www.farmaciasdirect.com/sitemap.xml"

# Delay entre peticiones
MIN_DELAY = 0.3
MAX_DELAY = 0.8

# Resultados por página (máximo probado: 100)
ROWS_PER_PAGE = 100

# Batch size para commits a la DB
BATCH_SIZE = 100


# ============================================================
# CONSULTA A EMPATHY.CO
# ============================================================
def consultar_empathy(query, rows=ROWS_PER_PAGE, start=0):
    """
    Hace una búsqueda en la API de Empathy.co de FarmaciasDirect.
    Retorna la respuesta JSON o None si falla.
    """
    params = {
        'query': query,
        'rows': str(rows),
        'start': str(start),
        'lang': 'es',
        'store': 'es',
        'instance': 'farmaciasdirect',
        'scope': 'desktop',
        'internal': 'true',
    }
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0',
        'Accept': 'application/json',
    }

    try:
        response = requests.get(EMPATHY_BASE_URL, params=params, headers=headers, timeout=30)
        if response.status_code != 200:
            logger.error(f"Error API Empathy: HTTP {response.status_code}")
            return None
        return response.json()
    except Exception as e:
        logger.error(f"Error en consulta Empathy: {e}")
        return None


# ============================================================
# DESCARGAR SITEMAP Y EXTRAER URLS DE PRODUCTOS
# ============================================================
def obtener_urls_sitemap():
    """
    Descarga el sitemap index de FarmaciasDirect y extrae todas las URLs
    de productos individuales de los sub-sitemaps.
    """
    logger.info("Descargando sitemap index...")
    headers = {'User-Agent': 'Mozilla/5.0'}

    try:
        r = requests.get(SITEMAP_INDEX_URL, headers=headers, timeout=30)
        r.raise_for_status()
    except Exception as e:
        logger.error(f"Error descargando sitemap: {e}")
        return []

    # Parsear sitemap index
    root = ElementTree.fromstring(r.content)
    ns = {'sm': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
    sitemap_urls = [loc.text for loc in root.findall('.//sm:loc', ns)
                    if 'products' in (loc.text or '')]

    logger.info(f"Encontrados {len(sitemap_urls)} sitemaps de productos")

    # Descargar cada sub-sitemap
    product_urls = []
    for i, url in enumerate(sitemap_urls):
        logger.info(f"Descargando sitemap {i+1}/{len(sitemap_urls)}...")
        try:
            r = requests.get(url, headers=headers, timeout=30)
            r.raise_for_status()
            sub_root = ElementTree.fromstring(r.content)
            urls = [loc.text for loc in sub_root.findall('.//sm:loc', ns)]
            product_urls.extend(urls)
            time.sleep(0.5)
        except Exception as e:
            logger.error(f"Error en sitemap {url}: {e}")
            continue

    logger.info(f"Total URLs de productos encontradas: {len(product_urls)}")
    return product_urls


def extraer_nombre_de_url(url):
    """
    Extrae un nombre de búsqueda a partir de la URL del producto.
    Ej: https://www.farmaciasdirect.es/products/paracetamol-normon-650mg -> "paracetamol normon 650mg"
    """
    # Obtener la última parte de la URL
    slug = url.rstrip('/').split('/')[-1]
    # Limpiar el slug
    nombre = slug.replace('-', ' ')
    # Quitar números muy largos que sean IDs
    nombre = re.sub(r'\b\d{10,}\b', '', nombre)
    return nombre.strip()


# ============================================================
# EXTRACCIÓN DE DATOS DE PRODUCTO
# ============================================================
def extraer_datos_producto(hit):
    """Extrae datos relevantes de un resultado de Empathy.co."""
    nombre = hit.get('__name', hit.get('nombre', 'Sin nombre'))
    url = hit.get('__url', hit.get('link', ''))
    marca = hit.get('marca', '')
    disponibilidad = hit.get('disponibilidad', '')
    en_stock = disponibilidad.lower() in ('disponible', 'in stock', 'in_stock', 'true', '1') if disponibilidad else True
    referencia = hit.get('referencia', hit.get('externalIdProducto', ''))

    # Precios
    precio = None
    precio_original = None

    prices = hit.get('__prices', {})
    if prices:
        try:
            current = prices.get('current', {})
            original = prices.get('original', {})
            if current:
                precio = Decimal(str(current.get('value', 0)))
            if original:
                precio_original = Decimal(str(original.get('value', 0)))
        except (InvalidOperation, ValueError):
            pass

    # Fallback a precioFinal/precioBase
    if precio is None:
        pf = hit.get('precioFinal')
        if pf:
            try:
                precio = Decimal(str(pf))
            except (InvalidOperation, ValueError):
                pass

    if precio_original is None:
        pb = hit.get('precioBase')
        if pb:
            try:
                precio_original = Decimal(str(pb))
            except (InvalidOperation, ValueError):
                pass

    # Categoría
    categoria = marca or ''

    # URL completa
    if url and not url.startswith('http'):
        url = f"https://www.farmaciasdirect.es{url}"

    return {
        'nombre': nombre,
        'url': url,
        'precio': precio,
        'precio_original': precio_original,
        'en_stock': en_stock,
        'categoria': categoria,
        'sku': str(referencia),
    }


# ============================================================
# SCRAPING VÍA BÚSQUEDA ALFABÉTICA
# ============================================================
def ejecutar_scraping(db, limit=0):
    """
    Extrae productos de FarmaciasDirect usando búsquedas alfabéticas
    por la API de Empathy.co.

    Estrategia: Buscar con combinaciones de 2 letras (aa, ab, ac...zz)
    para cubrir todos los productos del catálogo. Deduplicar por URL.
    """
    # Generar búsquedas: letras del abecedario + sílabas comunes en farmacia
    busquedas = []

    # Sílabas/prefijos comunes en productos de farmacia
    prefijos_farmacia = [
        'para', 'ibu', 'vita', 'crem', 'gel', 'ser', 'aceite',
        'champu', 'past', 'jar', 'gota', 'sol', 'pol', 'cap',
        'comp', 'spray', 'loci', 'poma', 'ung', 'sus',
        'anti', 'bio', 'pro', 'multi', 'omega', 'zinc', 'calc',
        'magn', 'hier', 'col', 'diet', 'prot', 'fib',
        'beb', 'inf', 'ped', 'adult', 'muj',
        'facial', 'corporal', 'capilar', 'solar', 'labial',
        'dental', 'bucal', 'nasal', 'ocular', 'aur',
    ]
    busquedas.extend(prefijos_farmacia)

    # Combinaciones de 2 letras para cubrir el resto
    import string
    for a in string.ascii_lowercase:
        for b in string.ascii_lowercase:
            combo = a + b
            if combo not in busquedas:
                busquedas.append(combo)

    urls_procesadas = set()
    exitos = 0
    errores = 0
    hoy = datetime.now(timezone.utc)

    # Primero, cargar URLs ya en la DB para no repetir
    productos_existentes = db.query(Producto.url).filter_by(farmacia=FARMACIA_NOMBRE).all()
    urls_existentes = {p[0] for p in productos_existentes}
    logger.info(f"Productos ya en DB: {len(urls_existentes)}")

    total_busquedas = len(busquedas)
    logger.info(f"Total de búsquedas a realizar: {total_busquedas}")

    for idx, query in enumerate(busquedas):
        if limit > 0 and exitos >= limit:
            break

        # Primera página para ver cuántos resultados hay
        resultado = consultar_empathy(query, rows=ROWS_PER_PAGE, start=0)
        if not resultado:
            continue

        catalog = resultado.get('catalog', {})
        num_found = catalog.get('numFound', 0)

        if num_found == 0:
            continue

        logger.info(f"[{idx+1}/{total_busquedas}] '{query}': {num_found} resultados")

        # Procesar todas las páginas de esta búsqueda
        total_pages = -(-num_found // ROWS_PER_PAGE)
        # Limitar a máx 10 páginas por búsqueda (1000 productos)
        total_pages = min(total_pages, 10)

        for page in range(total_pages):
            if limit > 0 and exitos >= limit:
                break

            if page > 0:
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))
                resultado = consultar_empathy(query, rows=ROWS_PER_PAGE, start=page * ROWS_PER_PAGE)
                if not resultado:
                    continue
                catalog = resultado.get('catalog', {})

            hits = catalog.get('content', [])
            if not hits:
                break

            for hit in hits:
                if limit > 0 and exitos >= limit:
                    break

                datos = extraer_datos_producto(hit)

                if not datos['url'] or datos['precio'] is None:
                    errores += 1
                    continue

                # Deduplicar
                if datos['url'] in urls_procesadas or datos['url'] in urls_existentes:
                    continue

                urls_procesadas.add(datos['url'])

                # Guardar en DB
                producto = db.query(Producto).filter_by(url=datos['url']).first()

                if not producto:
                    producto = Producto(
                        nombre=datos['nombre'],
                        url=datos['url'],
                        farmacia=FARMACIA_NOMBRE,
                        categoria=datos['categoria'],
                        ean=datos['sku'],
                    )
                    db.add(producto)
                    db.flush()
                else:
                    producto.nombre = datos['nombre']
                    producto.categoria = datos['categoria']

                precio_record = Precio(
                    producto_id=producto.id,
                    precio=datos['precio'],
                    precio_original=datos['precio_original'],
                    en_stock=datos['en_stock'],
                    fecha_captura=hoy,
                )
                db.add(precio_record)
                exitos += 1

                if exitos % BATCH_SIZE == 0:
                    db.commit()
                    logger.info(f"  Guardados: {exitos} productos nuevos")

        # Delay entre búsquedas
        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    # Commit final
    db.commit()

    logger.info("=" * 60)
    logger.info("RESUMEN FINAL")
    logger.info("=" * 60)
    logger.info(f"Productos nuevos extraídos: {exitos}")
    logger.info(f"Productos sin precio/URL: {errores}")
    logger.info(f"Total en DB: {db.query(Producto).filter_by(farmacia=FARMACIA_NOMBRE).count()}")


# ============================================================
# EXPORTAR A EXCEL
# ============================================================
def exportar_a_excel(db, filename=None):
    if filename is None:
        export_dir = os.path.join(PROJECT_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        filename = os.path.join(export_dir, 'precios_farmaciasdirect.xlsx')

    """Exporta los precios más recientes a Excel con formato profesional."""
    from sqlalchemy import text
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = text("""
        SELECT
            p.nombre,
            p.url,
            p.categoria,
            p.ean,
            pr.precio,
            pr.precio_original,
            pr.en_stock,
            pr.fecha_captura
        FROM productos p
        JOIN precios pr ON p.id = pr.producto_id
        WHERE p.farmacia = :farmacia
        AND pr.id = (
            SELECT pr2.id FROM precios pr2
            WHERE pr2.producto_id = p.id
            ORDER BY pr2.fecha_captura DESC
            LIMIT 1
        )
        ORDER BY p.categoria, p.nombre
    """)

    result = db.execute(query, {"farmacia": FARMACIA_NOMBRE})
    rows = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Precios FarmaciasDirect"

    # Estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_par = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_stock_si = Font(name='Calibri', color='2E7D32')
    fill_stock_no = Font(name='Calibri', color='C62828')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    headers = ['Farmacia', 'Nombre', 'Marca', 'Referencia', 'Precio (€)', 'Precio Original (€)', 'Descuento', 'En Stock', 'URL', 'Fecha']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        nombre = row[0]
        url = row[1]
        marca = row[2]
        ref = row[3]
        precio = float(row[4]) if row[4] else 0
        precio_original = float(row[5]) if row[5] else None
        en_stock = row[6]
        fecha = row[7]

        descuento = ""
        if precio_original and precio_original > 0 and precio < precio_original:
            pct = ((precio_original - precio) / precio_original) * 100
            descuento = f"-{pct:.0f}%"

        fill = fill_par if row_idx % 2 == 0 else fill_impar
        data = [
            FARMACIA_NOMBRE, nombre, marca, ref or '',
            precio, precio_original, descuento,
            'Sí' if en_stock else 'No', url,
            fecha.strftime('%Y-%m-%d %H:%M') if fecha else '',
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = align_left

        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=5).alignment = align_right
        if precio_original:
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00 €'
        ws.cell(row=row_idx, column=6).alignment = align_right
        ws.cell(row=row_idx, column=7).alignment = align_center
        ws.cell(row=row_idx, column=7).font = Font(name='Calibri', color='C62828', bold=True)

        stock_cell = ws.cell(row=row_idx, column=8)
        stock_cell.alignment = align_center
        stock_cell.font = fill_stock_si if en_stock else fill_stock_no

        ws.cell(row=row_idx, column=10).alignment = align_center

    column_widths = {1: 16, 2: 55, 3: 20, 4: 16, 5: 14, 6: 18, 7: 12, 8: 11, 9: 60, 10: 18}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    wb.save(filename)
    logger.info(f"Exportados {len(rows)} productos a '{filename}'")
    return len(rows)


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Scraper de precios FarmaciasDirect (vía Empathy.co API)")
    parser.add_argument("--limit", type=int, default=0, help="Limitar a N productos (0=todos)")
    parser.add_argument("--export", action="store_true", help="Solo exportar datos existentes a Excel")
    parser.add_argument("--output", type=str, default=None, help="Ruta del archivo de salida")
    args = parser.parse_args()

    engine = get_engine()
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    try:
        if args.export:
            logger.info("Exportando datos existentes a Excel...")
            exportar_a_excel(db, filename=args.output)
        else:
            logger.info("=" * 60)
            logger.info("SCRAPING DE FARMACIASDIRECT VÍA EMPATHY.CO API")
            logger.info("=" * 60)
            ejecutar_scraping(db, limit=args.limit)

            logger.info("")
            logger.info("=" * 60)
            logger.info("EXPORTANDO A EXCEL")
            logger.info("=" * 60)
            exportar_a_excel(db, filename=args.output)
    finally:
        db.close()


if __name__ == "__main__":
    main()
